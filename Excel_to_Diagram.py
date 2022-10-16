# import ctypes
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter.ttk import Combobox
import numpy
import pandas as pd
from matplotlib import pyplot as plt
import openpyxl
import re


class Chart:

    def Pie_Chart(self):  # 输入字典dic
        a = []
        b = []
        ratio = []
        for k in self:
            a.append(k)
            b.append(self[k])

        count = sum(b)  # 总个数
        label = a

        for u in b:
            i = u / count  # 算出每一个patient的使用比例%
            ratio.append(i)
            plt.plot(ratio)

        plt.title(entry_title.get(), fontsize='xx-large', fontweight='heavy')  # 饼图标题
        plt.pie(ratio,  # 数据
                labels=label,  # 数据的标签
                # explode=explode,  # 某些部分突出显示, 爆炸式
                pctdistance=0.85,
                autopct='%.1f%%',  # 百分百格式, 保留一位小数
                # shadow=True,  # 底部添加阴影
                radius=1.2,  # 饼图半径
                wedgeprops={'linewidth': 1, 'edgecolor': 'white'},  # 内外边界的属性
                textprops={'fontsize': 12, 'color': 'black'},  # 标签文字的属性
                startangle=90  # 饼图初始角度
                )
        plt.show()

    def Bar_Chart(self):  # 输入字典dic
        k = []
        v = []
        for kkk in self:
            k.append(kkk)
            v.append(self[kkk])

        fig, ax = plt.subplots()

        b = ax.bar(k, v)

        for a, b in zip(k, v):
            ax.text(a, b, b, ha='center', va='bottom')  # 在柱顶端标上数字

        plt.title(entry_title.get(), fontsize='xx-large', fontweight='heavy')  # 标题
        plt.xlabel(entry_x.get(), fontsize='large', fontweight='heavy')  # x轴标签
        plt.ylabel(entry_y.get(), fontsize='large', fontweight='heavy')  # y轴标签
        # plt.xticks(rotation=310)  # x轴标签旋转
        fig.autofmt_xdate()
        plt.grid(True, linestyle='-.', axis='y', alpha=0.5)  # 添加网格线
        plt.show()

    def Radial_Col_Chart(self):  # 输入字典dic
        # dic_sorted = sorted(self.items(), key=lambda x: x[1], reverse=False)
        # key是排序的依据, 按元素第二个字段排序, 也就是字典的value ; reverse为F, 从小到大排序
        # print(dic_sorted)

        df = pd.DataFrame(pd.Series(self), columns=['Number'])
        df = df.reset_index().rename(columns={'index': 'Patient_id'})
        df = df.sort_values(by=['Number'])
        print(df)

        # 创建画布
        plt.figure(figsize=(10, 5))
        ax = plt.subplot(111, polar=True)
        plt.axis('off')

        upperLimit = 800
        lowerLimit = 1
        labelPadding = 4

        max_value = df['Number'].max()  # 拿到最大值
        slope = (max_value - lowerLimit) / max_value
        heights = slope * df.Number + lowerLimit

        # 计算条形图的宽度
        width = 2 * numpy.pi / len(df.index)

        # 计算角度
        indexes = list(range(1, len(df.index) + 1))
        angles = [element * width for element in indexes]

        # 绘制条形图
        bars = ax.bar(
            x=angles,
            height=heights,
            width=width,
            bottom=lowerLimit,
            linewidth=2,
            edgecolor="white",
            color="#61a4b2",
        )

        # 添加标签
        for bar, angle, height, label in zip(bars, angles, heights, df['Patient_id']):

            # 旋转
            rotation = numpy.rad2deg(angle)

            # 翻转
            alignment = ""
            if numpy.pi / 2 <= angle < 3 * numpy.pi / 2:
                alignment = "right"
                rotation = rotation + 180
            else:
                alignment = "left"

            # 最后添加标签
            ax.text(
                x=angle,
                y=lowerLimit + bar.get_height() + labelPadding,
                s=label,
                ha=alignment,
                va='center',
                rotation=rotation,
                rotation_mode="anchor")

        plt.show()

    def H_Bar_Chart(self):
        k = []
        v = []
        for kkk in self:
            k.append(kkk)
            v.append(self[kkk])

        # 初始化画布和坐标系数据
        fig, ax = plt.subplots()
        b = ax.barh(range(len(k)), v,
                    color=['cornflowerblue', 'lightcoral', 'peru', 'mediumseagreen', 'orchid', 'tan', 'plum',
                           'darkgrey', 'steelblue', 'turquoise'])

        # 在柱的顶端标上数据
        for label in b:
            width = label.get_width()
            ax.text(width, label.get_y() + label.get_height() / 2, '%d' % int(width), ha='left', va='center')

        # 设置y轴,标上patient
        ax.set_yticks(range(len(k)))
        ax.set_yticklabels(k)

        # 添加网格线和标题
        plt.grid(True, linestyle='dashdot', axis='x', alpha=0.5)
        plt.title(entry_title.get(), loc='center', fontsize='25', fontweight='bold')  # 标题
        plt.xlabel(entry_x.get(), fontsize='large', fontweight='heavy')  # x轴标签
        plt.ylabel(entry_y.get(), fontsize='large', fontweight='heavy')  # y轴标签
        plt.show()


def doProcess():
    if selected_radio.get() == 1:
        Chart.Pie_Chart(readExcel())
    elif selected_radio.get() == 2:
        Chart.Bar_Chart(readExcel())
    elif selected_radio.get() == 3:
        Chart.H_Bar_Chart(readExcel())
    else:
        tk.messagebox.showinfo('提示', '请选择图表类型')


def selectExcel():
    selected_file.set('')
    file_name = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx')])

    text1.delete(0, END)  # entry里选择第二个的时候, 删除前一个
    text1.insert(INSERT, file_name)

    if file_name:
        selected_file.set(file_name)


def closeThisWindow():
    if messagebox.askokcancel("退出", "确定要退出吗?"):
        root.destroy()


def readExcel():
    # tk.messagebox.showinfo('提示', '处理Excel文件的示例程序。')
    cells_list.clear()
    PD_list.clear()
    path = selected_file.get()
    workbook = openpyxl.load_workbook(path)  # 获取文件

    sheet_names = workbook.sheetnames  # 获取所有表的名字
    sheet = workbook.active  # 获取活跃表对象
    row_num = sheet.max_row  # 获取表中最大的行数
    col_num = sheet.max_column  # 获取表中最大的列数

    for col in range(1, col_num + 1):  # 获取表中所有单元格 放入list
        for row in range(1, row_num + 1):
            cell = sheet.cell(row, col)
            if cell.value is not None:
                cells_list.append(cell.value)
    patient_names = ("4DLung",
                     "30x30x30",
                     "0505new",
                     "0505new2",
                     "002441",
                     "002443",
                     "002443TRN",
                     "002445",
                     "002445TRN",
                     "123456AA",
                     "141414",
                     "306941",
                     "0009999012",
                     "0010010777",
                     "0010117535",
                     "20140130",
                     "20160783",
                     "20190814B",
                     "111222333",
                     "111222333_NoContours",
                     "333222111",
                     "666555444",
                     "999888777",
                     "AAMDRSS2016PS",
                     "AdaptCBCT",
                     "AdaptCBCT2",
                     "AdaptCTDemo",
                     "ADTProstate",
                     "ANON76497",
                     "ArcCheck",
                     "ATP13",
                     "AutoCP",
                     "AutoMargin51102",
                     "BigAlFFS",
                     "BigAlHFS",
                     "BigPhantom",
                     "Bill",
                     "BOLUS11",
                     "Brain5TargetsTRN",
                     "breast1",
                     "breast1IMRT",
                     "BreastTRN",
                     "c1htth",
                     "CarbonLung",
                     "cem0tgdznno",
                     "CHESTnonPHANTOM",
                     "closedCentralL",
                     "cn1xt",
                     "CNNS1",
                     "cPatient",
                     "cPatient.AlgTest",
                     "cPatientProton",
                     "CrossPhantom",
                     "CT1",
                     "ctMRdosePhan",
                     "CTMRX30",
                     "CTMRXM20Y30Z50",
                     "CTMRYM40",
                     "CTMRZM50",
                     "CTSdemoHN",
                     "CTSdemoProst",
                     "CTSlung1",
                     "CTSprostate1",
                     "CTwithCTandPET",
                     "CUBE120kV",
                     "cwkjc5m51",
                     "cxrrbvw3",
                     "DemoPatient11",
                     "DepthSSDtest",
                     "DEVIL",
                     "DVHPatient1",
                     "DVHStats",
                     "DynMinLeafGap",
                     "EKK999999",
                     "ePHANTOM",
                     "FFP0CHEST",
                     "FFP88811X",
                     "FFPabdo",
                     "FFPDCMORIGIN",
                     "FFPEDFeetToHea",
                     "FFPEDHeadToFee",
                     "FFS15",
                     "FFS88811X",
                     "FFSDCMORIGIN",
                     "FFSEDFeetToHea",
                     "FFSEDHeadToFee",
                     "FFSPELVIS",
                     "ForceEDpat",
                     "ForceFillHN",
                     "FrozenDose4",
                     "FrozenTest",
                     "FrozenTest2",
                     "FrozenTest3",
                     "FrozenTestProton",
                     "FusionProstate",
                     "HalBreastXiO",
                     "HeadNeck_air",
                     "HeadProne",
                     "HFP00HEAD",
                     "HFP0CHEST",
                     "HFP88811X",
                     "HFPDCMORIGIN",
                     "HFPEDFeetToHea",
                     "HFPEDHeadToFee",
                     "HFPPELVIS",
                     "HFS00HEAD",
                     "HFS0CHEST",
                     "HFS30",
                     "HFS88811X",
                     "HFSDCMORIGIN",
                     "HFSEDFeetToHea",
                     "HFSEDHeadToFee",
                     "HFSPELVIS",
                     "HN",
                     "HN1",
                     "HNBig",
                     "HNnonPHANTOM",
                     "HNrtogIMRT",
                     "IMPTcase1",
                     "inhomtest",
                     "InvliChecksum",
                     "ionPINRTFP",
                     "ionPINRTFS",
                     "ionPINRTHP",
                     "ionPINRTHS",
                     "JamesCookMonac",
                     "K5dQtHx29P",
                     "larnumslices",
                     "Liver",
                     "LLungSBRT",
                     "LucyPhantom",
                     "Lung_air",
                     "lung_air_1",
                     "M511SSR",
                     "Matrixx",
                     "me010666x",
                     "monCOUCH",
                     "monDRPProstate",
                     "monDVHProstate",
                     "MonFP15B",
                     "monFP20B",
                     "monFP25B",
                     "monFusionProst",
                     "monFusionProstSAVE",
                     "monHN",
                     "monHPQC3121",
                     "monIMRTslabPhantm",
                     "monLung",
                     "monMLCBrain",
                     "monMLCProstate",
                     "monOROPHARYNX",
                     "monPhantom",
                     "monPlans",
                     "monPROSTATE",
                     "monPROSTATEexport",
                     "monPROSTATEmulti",
                     "monPROSTATEOLD",
                     "monPROSTATEproton",
                     "monSpherePhan",
                     "MRcntrSCO",
                     "MRIAdaptFFP",
                     "MRIAdaptFFS",
                     "MRIAdaptHFP",
                     "MRIAdaptHFS",
                     "MRIAdaptOrient",
                     "MRIDemo",
                     "MRITriangles",
                     "MRLCouch",
                     "MRX1Y1Z1",
                     "MRX20YM30Z40",
                     "MRX60",
                     "MRXM65Y55Z60",
                     "MRY50",
                     "MRZM60",
                     "MultipleStructures",
                     "nonphilips",
                     "OcuPro",
                     "OROPHARYNX",
                     "P20200925",
                     "PATIMG01",
                     "PHANTOM",
                     "PhShiftFFP50",
                     "PhShiftFFS50",
                     "PhShiftHFP50",
                     "PhShiftHFS50",
                     "PlanQuality",
                     "PointReg",
                     "PRETTYPROTONS",
                     "PROFFS18",
                     "PROHFP15",
                     "PROHFS13",
                     "prostate",
                     "Prostate_air",
                     "Prostate_norm",
                     "ProstateNodes",
                     "ProstShiftSAM",
                     "R95689791",
                     "referencePhant",
                     "RFC33982",
                     "RFC34976",
                     "RFC35090",
                     "roll",
                     "ROQS25",
                     "SBRTLung",
                     "SBRTSpine",
                     "SFUD_2",
                     "SFUD_3",
                     "SFUD1",
                     "smokePATIENT",
                     "smokePATIENT2",
                     "SPCIMGTEST",
                     "SpotLevel",
                     "SPTOriginalEUR",
                     "SSDTest2",
                     "Test1Proton",
                     "Test2Proton",
                     "Test9Proton",
                     "ThoracicCTPET",
                     "TolBreast",
                     "TPPC17A03PH",
                     "TwoBolus",
                     "UnityPatient1",
                     "UPMCLIV2",
                     "UPMCPANC1",
                     "UserAuth",
                     "VarianVMAT",
                     "wccccm",
                     "WeightLoss",
                     "XiOLung",
                     "EHU1412",
                     "4DCIRS1",
                     "111111",
                     "0000463294.anonymized",
                     "0000492064.anonymized",
                     "42482527",
                     "42482527_PerformanceTest",
                     "77378376",
                     "A4DCIRS1",
                     "cPatientCTRS_NonSquare",
                     "DUCKDAISY",
                     "HeadNeck_air_PerformanceTest",
                     "HP6834",
                     "Liver-4537",
                     "LockedLiver",
                     "lung_air_PerfomanceTest",
                     "MRIDemo177",
                     "MRIDemoCP",
                     "PatientLiver",
                     "ProtonCase",
                     "ProtonHBig",
                     "ProtonHSmall",
                     "ProtonPlanReview",
                     "ROQS25",
                     "smokePATIENT",
                     "smokePATIENT2",
                     "Test1",
                     "Test1IDD",
                     "Test3",
                     "Test3Air",
                     "Test4",
                     "Test4Water",
                     "Test6",
                     "Test8IDD",
                     "Test10",
                     "Test11RBE")
    dicom_names = ("1ATLABDO_FFPMR_Day1",
                   "1ATLABDO_FFPMR_Day2",
                   "1ATLABDO_FFSMR_Day1",
                   "1ATLABDO_FFSMR_Day2",
                   "1ATLABDO_HFPMR_Day1",
                   "1ATLABDO_HFPMR_Day2",
                   "1ATLABDO_HFSMR_Day1",
                   "1ATLABDO_HFSMR_Day2",
                   "3D",
                   "3DSTACarc",
                   "4DCIRS1",
                   "4DLung",
                   "4011",
                   "393456",
                   "17112006",
                   "6012247133",
                   "AddonMLC",
                   "Applicator",
                   "BigAlFFS",
                   "BigAlHFS",
                   "Bill",
                   "BlackPixels",
                   "Block",
                   "BO2",
                   "BolusImportLiver",
                   "BolusOncentra2",
                   "BREAST0719",
                   "BUG_6255_PhantomPlan",
                   "Carbon_FFP",
                   "Carbon_FFP_T_PTS",
                   "Carbon_FFS",
                   "Carbon_FFS_T_PTS",
                   "Carbon_HFP",
                   "Carbon_HFP_T_PTS",
                   "Carbon_HFS",
                   "Carbon_HFS_T_PTS",
                   "ch2o1xr2rmFFP",
                   "ch2o1xr2rmFFS",
                   "ch2o1xr2rmHFP",
                   "ch2o1xr2rmHFS",
                   "Chordoma_2",
                   "ConformalRT",
                   "cPatientNonSquarePixelMR",
                   "CS_CT_iBEAM_CouchTop",
                   "CS_CT_MULTISERIES",
                   "CS_CTMR_PICKER",
                   "CS_MR_TILT",
                   "CS_MRI_NON_SQR_PX",
                   "CS_NEARLYAXIALMR",
                   "CS_NOMERG_SIEMENS",
                   "CS_PET_BLACK_HOLES",
                   "CS_PET_GESLICE",
                   "CS_PET_GEVOLUME",
                   "CS_PETCT_GEDISC_JUN2002",
                   "CS_RTD_ECLIPSE",
                   "CS_RTD_ECLIPSE_IMPORT",
                   "CS_RTD_KONRAD",
                   "CS_RTD_ONCENTRAsep2005",
                   "CS_RTD_TOMOHN",
                   "CS_RTD_XIOIMRTARC",
                   "CS_RTD_XIORTPLINK",
                   "CS_RTP_ACQSIM_APRIL2005",
                   "CS_RTP_ADVSIM_JUL2005",
                   "CS_RTP_ONCENTRA_JUL2005",
                   "CS_RTP_THERAPLAN",
                   "CS_RTP_XIO_FORSSIMPORT",
                   "CS_RTS_ACQSIM_NASO",
                   "CS_RTS_COHERENCE",
                   "CT_RTS_COHERENCE",
                   "CT_UnEven",
                   "CT_VF",
                   "CTMR",
                   "CTOR03",
                   "CTPETSiemens",
                   "CTSdemoProst",
                   "DCAT_CDR_Bug_Data",
                   "Demo_Phantom_1",
                   "DemoPatient10",
                   "dMLC",
                   "DRPDepthTest",
                   "DuplicateInstanceNumber",
                   "DuplicateSliceLocation",
                   "DuplicateSliceLocationNoInstanceNumber",
                   "DynConformalArc",
                   "ED_TestPhantoms",
                   "EnergyError",
                   "ErrorThrowout",
                   "FEET1st^SUPINE_ABC-DEFGHIJKLMNOPQRSTUVWXYZ.abcdefghijklmnopqrstu",
                   "FFP_PhantomSkull",
                   "FFPabdo",
                   "FFS",
                   "FFS_RealPatient",
                   "FFS_roll-15",
                   "Fieldsplit",
                   "FusionProstate",
                   "GE_AdvSim_2field_BreastPlan",
                   "GEPatient",
                   "GESIM",
                   "Grid9",
                   "HeadSupine",
                   "HFP",
                   "HFP_1024Images",
                   "HFP_roll-15",
                   "HFS_RealPatientRTSS",
                   "HFS_roll-15",
                   "InvalidSnout",
                   "InvalidSP",
                   "ionRSThicknessError",
                   "IOPnoTranswithoutMacro",
                   "IOPxMACRO",
                   "Liver",
                   "mARC",
                   "MMRO17A02EL_REG_Nucletron",
                   "MMRO17A06EL_REG_Varian",
                   "MMRO17A06RA_REG",
                   "MMRO17A06RA_REG_RaySearch",
                   "MMRO17A10VA_REG_3SET",
                   "MMRO19A03EL",
                   "Mon4TESTING01",
                   "MONACODVH",
                   "MonIonDicom",
                   "monPhan",
                   "MR",
                   "MR_OBLIQUE_DATA",
                   "MR_PLANE_CUTS",
                   "MRIAdaptFFP",
                   "MRIAdaptFFS",
                   "MRIAdaptHFP",
                   "MRIAdaptHFP_IPP",
                   "MRIAdaptHFS",
                   "MRIAdaptHFS_IPP",
                   "MRMultiPlan",
                   "MRPuppetOrient",
                   "MRTestDemo",
                   "MRTestPatientHN",
                   "MultiDelivery",
                   "MultiDuplicateSliceLocation",
                   "MultiFrac",
                   "NomappingTU",
                   "Oncentra3",
                   "OrientationTestPatient1",
                   "p5MUeeYS0w",
                   "PatientLiver",
                   "Ph.Orig.FFP",
                   "Ph.Orig.FFS",
                   "Ph.Orig.HFP",
                   "Ph.Orig.HFS",
                   "Ph.PrismFadeZ",
                   "Ph.Shift.FFP",
                   "Ph.Shift.FFS",
                   "Ph.Shift.HFP",
                   "Ph.Shift.HFS",
                   "Ph.Shift.HFS_Simplify",
                   "Phantom_NullStudy",
                   "PhDosePrism10",
                   "PhilipsPatient",
                   "PhillipsPetCT",
                   "ProneData",
                   "prostate-umb1-CBCT",
                   "prostate-umb1-MRI",
                   "ProstShift_DICOM",
                   "ProtonHSmall",
                   "Regression_5.5",
                   "RFC27418Dcm",
                   "RolledCT",
                   "RSIDerror",
                   "SE000001anon",
                   "secondarymonPROSTATE",
                   "secondaryMRIDemo",
                   "Segerror",
                   "SetupBeam",
                   "SFUD4_ClinicalCase",
                   "SiemensPatient",
                   "SnoutIDError",
                   "SpecExprt",
                   "SpecImg",
                   "SPHERE120kV",
                   "SSDtest",
                   "SSDtest2",
                   "stepnshoot",
                   "SwedishDicom",
                   "T1Prostate",
                   "TestProstatePatient",
                   "TOSHIBA_IMG",
                   "TPPC17A03BR_REG_MIM",
                   "TPPC17A04BR_REG_Brainlab",
                   "TuneIDerror",
                   "Unity01",
                   "vmat",
                   "vmatnew",
                   "VSADtest",
                   "Wedge",
                   "WinLevelPET",
                   "YawPitchRollCT",
                   "ZeroSliceThickness",
                   "ZZZ_CT_1.25mmThinkness",
                   "4DCIRS1SRO",
                   "4DElektaXVi",
                   "4DElektaXviNoStudy",
                   "4DGE1",
                   "4DGE2",
                   "4DGE4",
                   "4DGE5",
                   "4DGE6",
                   "4DManyStudies",
                   "4DPhilips1",
                   "4DPhilips2",
                   "4DPhilips3",
                   "4DPhilips4",
                   "4DPhilips5",
                   "4DPhilips6",
                   "4DPhilips7",
                   "4DPhilips8",
                   "4DSiemens1",
                   "4DSiemens2",
                   "4DSiemens3",
                   "4DSiemens4",
                   "4DSiemens5",
                   "4DSiemens6",
                   "4DToshiba4",
                   "000006+1485",
                   "000006+1541",
                   "000006+1648",
                   "0000061485",
                   "CarpePTS",
                   "crjcr1q",
                   "CTdisney",
                   "ctMRdosePhan_MRI",
                   "ctMRdosePhanMRsurvey",
                   "DataBug15",
                   "DosePhantom",
                   "DosePhantoms",
                   "DuplicateSliceLocation4D",
                   "ExcCase1",
                   "I18Neiou",
                   "IPPShift",
                   "Multiple_Size_Image",
                   "PET_Phantom",
                   "PETSUVA",
                   "PETSUVB",
                   "phantom",
                   "SamplePatientsForImport",
                   "testSUV",
                   "WINLEVEL_PRESET_3",
                   "WINLEVEL_PRESET_4")

    for c in cells_list:
        for patient in patient_names:
            if re.search(r'\b%s\b' % patient, str(c), re.IGNORECASE):
                PD_list.append(patient)

        for dicom in dicom_names:
            if re.search(r'\b%s\b' % dicom, str(c), re.IGNORECASE):
                PD_list.append(dicom)

    # 对list进行筛选
    total = {}  # 字典 包含所有的id和它的数量
    for item in PD_list:  # 找到每个patient_id的使用次数
        if item in total.keys():
            total[item] += 1
        else:
            total[item] = 1

    if filter_num != 0:
        new_filter_num = int(combo_num.get())
        if combo_symbol.get() == '>=':
            changed = {k: v for k, v in total.items() if v >= int(new_filter_num)}
            return changed
        elif combo_symbol.get() == '>':
            changed = {k: v for k, v in total.items() if v > int(new_filter_num)}
            return changed
        elif combo_symbol.get() == '<':
            changed = {k: v for k, v in total.items() if v < int(new_filter_num)}
            return changed
        elif combo_symbol.get() == '<=':
            changed = {k: v for k, v in total.items() if v <= int(new_filter_num)}
            return changed
    return total

    # test_patient = workbook.sheet_by_index(0)  # 拿到第一个sheet
    # test_dicom = workbook.sheet_by_index(1)
    #
    # test_patient_data = test_patient.col_values(1)[0:2028]  # 拿到第一个sheet里的第2列
    # test_dicom_data = test_dicom.col_values(1)[0:501]  # patients
    # if selected_radio.get() == 1:
    #     Chart.Pie_Chart(read_excel(test_patient_data))
    # elif selected_radio.get() == 2:
    #     Chart.Bar_Chart(read_excel(test_patient_data))
    # elif selected_radio.get() == 3:
    #     Chart.H_Bar_Chart(read_excel(test_patient_data))


if __name__ == "__main__":
    # style = Style(theme = "cosmo")

    root = tk.Tk()
    # ctypes.windll.shcore.SetProcessDpiAwareness(1)
    # ScaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0)  # 调用api获得当前的缩放因子
    # root.tk.call('tk', 'scaling', ScaleFactor / 90)  # 设置缩放因子
    cells_list = []
    PD_list = []
    root.title("PD_Diagram_Maker V1.0")
    root.geometry("500x300+470+200")  # 窗口初始化大小和位置

    selected_radio = IntVar()
    selected_file = StringVar()
    filter_var = IntVar()
    filter_var.set('>=')
    filter_num = IntVar()
    filter_num.set('0')

    label1 = Label(root, text="请选择文件:")
    text1 = Entry(root, background="blue", width=42)
    button1 = Button(root, text="浏览", width=8, command=selectExcel)
    button2 = Button(root, text="作图", width=8, command=doProcess)
    button3 = Button(root, text="退出", width=8, command=closeThisWindow)
    r1 = Radiobutton(root, text="饼图", value=1, variable=selected_radio)
    r2 = Radiobutton(root, text="柱状图", value=2, variable=selected_radio)
    r3 = Radiobutton(root, text="横条图", value=3, variable=selected_radio)
    label_filter = Label(root, text="过滤条件:")
    combo_symbol = Combobox(root, textvariable=filter_var, width=3, values=['>', '>=', '<', '<='])
    combo_num = Entry(root, textvariable=filter_num, width=3)
    label_title = Label(root, text="图表标题:")
    label_x = Label(root, text="X轴名称:")
    label_y = Label(root, text="Y轴名称:")
    entry_title = Entry(root, width=20)
    entry_x = Entry(root, width=20)
    entry_y = Entry(root, width=20)

    # <editor-fold desc="UI位置参数">
    label1.pack()  # 调用pack方法将label标签显示在主界面
    text1.pack()
    button1.pack()
    button2.pack()
    button3.pack()
    r1.pack()
    r2.pack()
    r3.pack()
    label_filter.pack()
    combo_symbol.pack()
    combo_num.pack()
    label_title.pack()
    label_x.pack()
    label_y.pack()
    entry_title.pack()
    entry_x.pack()
    entry_y.pack()

    label1.place(x=25, y=30)
    text1.place(x=100, y=30)
    button1.place(x=400, y=28)
    button2.place(x=160, y=240)
    button3.place(x=260, y=240)
    r1.place(x=120, y=60)
    r2.place(x=210, y=60)
    r3.place(x=300, y=60)

    label_filter.place(x=130, y=100)
    combo_symbol.place(x=200, y=100)
    combo_num.place(x=250, y=100)

    label_title.place(x=130, y=130)
    label_x.place(x=130, y=160)
    label_y.place(x=130, y=190)
    entry_title.place(x=200, y=130)
    entry_x.place(x=200, y=160)
    entry_y.place(x=200, y=190)
    # </editor-fold>

    root.mainloop()

