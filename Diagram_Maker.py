import ctypes
import os
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
# import tkinter.ttk as ttk
# from tkinter.ttk import Combobox
import numpy
import pandas as pd
from matplotlib import pyplot as plt

import openpyxl
from openpyxl.styles import PatternFill
import re

from openpyxl import Workbook


class Chart:

    def Pie_Chart(self):  # 输入字典dic
        a = []
        b = []
        ratio = []
        for k in self:
            a.append(k)
            b.append(self[k])

        count = sum(b)  # 总个数
        label = a

        for u in b:
            i = u / count  # 算出每一个patient的使用比例%
            ratio.append(i)
            plt.plot(ratio)

        plt.title(entry_diagram_title.get(), fontsize='xx-large', fontweight='heavy')  # 饼图标题
        plt.pie(ratio,  # 数据
                labels=label,  # 数据的标签
                # explode=explode,  # 某些部分突出显示, 爆炸式
                pctdistance=0.85,
                autopct='%.1f%%',  # 百分百格式, 保留一位小数
                # shadow=True,  # 底部添加阴影
                radius=1.2,  # 饼图半径
                wedgeprops={'linewidth': 1, 'edgecolor': 'white'},  # 内外边界的属性
                textprops={'fontsize': 12, 'color': 'black'},  # 标签文字的属性
                startangle=90  # 饼图初始角度
                )
        plt.show()

    def Bar_Chart(self):  # 输入字典dic
        k = []
        v = []
        for kkk in self:
            k.append(kkk)
            v.append(self[kkk])

        fig, ax = plt.subplots()

        b = ax.bar(k, v)

        for a, b in zip(k, v):
            ax.text(a, b, b, ha='center', va='bottom')  # 在柱顶端标上数字

        plt.title(entry_diagram_title.get(), fontsize='xx-large', fontweight='heavy')  # 标题
        plt.xlabel(entry_x.get(), fontsize='large', fontweight='heavy')  # x轴标签
        plt.ylabel(entry_y.get(), fontsize='large', fontweight='heavy')  # y轴标签
        # plt.xticks(rotation=310)  # x轴标签旋转
        fig.autofmt_xdate()
        plt.grid(True, linestyle='-.', axis='y', alpha=0.5)  # 添加网格线
        plt.show()

    def Radial_Col_Chart(self):  # 输入字典dic
        # dic_sorted = sorted(self.items(), key=lambda x: x[1], reverse=False)
        # key是排序的依据, 按元素第二个字段排序, 也就是字典的value ; reverse为F, 从小到大排序
        # print(dic_sorted)

        df = pd.DataFrame(pd.Series(self), columns=['Number'])
        df = df.reset_index().rename(columns={'index': 'Patient_id'})
        df = df.sort_values(by=['Number'])
        print(df)

        # 创建画布
        plt.figure(figsize=(10, 5))
        ax = plt.subplot(111, polar=True)
        plt.axis('off')

        upperLimit = 800
        lowerLimit = 1
        labelPadding = 4

        max_value = df['Number'].max()  # 拿到最大值
        slope = (max_value - lowerLimit) / max_value
        heights = slope * df.Number + lowerLimit

        # 计算条形图的宽度
        width = 2 * numpy.pi / len(df.index)

        # 计算角度
        indexes = list(range(1, len(df.index) + 1))
        angles = [element * width for element in indexes]

        # 绘制条形图
        bars = ax.bar(
            x=angles,
            height=heights,
            width=width,
            bottom=lowerLimit,
            linewidth=2,
            edgecolor="white",
            color="#61a4b2",
        )

        # 添加标签
        for bar, angle, height, label in zip(bars, angles, heights, df['Patient_id']):

            # 旋转
            rotation = numpy.rad2deg(angle)

            # 翻转
            alignment = ""
            if numpy.pi / 2 <= angle < 3 * numpy.pi / 2:
                alignment = "right"
                rotation = rotation + 180
            else:
                alignment = "left"

            # 最后添加标签
            ax.text(
                x=angle,
                y=lowerLimit + bar.get_height() + labelPadding,
                s=label,
                ha=alignment,
                va='center',
                rotation=rotation,
                rotation_mode="anchor")

        plt.show()

    def H_Bar_Chart(self):
        k = []
        v = []
        for kkk in self:
            k.append(kkk)
            v.append(self[kkk])

        # 初始化画布和坐标系数据
        fig, ax = plt.subplots()
        b = ax.barh(range(len(k)), v,
                    color=['cornflowerblue', 'lightcoral', 'peru', 'mediumseagreen', 'orchid', 'tan', 'plum',
                           'darkgrey', 'steelblue', 'turquoise'])

        # 在柱的顶端标上数据
        for label in b:
            width = label.get_width()
            ax.text(width, label.get_y() + label.get_height() / 2, '%d' % int(width), ha='left', va='center')

        # 设置y轴,标上patient
        ax.set_yticks(range(len(k)))
        ax.set_yticklabels(k)

        # 添加网格线和标题
        plt.grid(True, linestyle='dashdot', axis='x', alpha=0.5)
        plt.title(entry_diagram_title.get(), loc='center', fontsize='25', fontweight='bold')  # 标题
        plt.xlabel(entry_x.get(), fontsize='large', fontweight='heavy')  # x轴标签
        plt.ylabel(entry_y.get(), fontsize='large', fontweight='heavy')  # y轴标签
        plt.show()


def txt_or_cs(file):
    if file.endswith('.txt'):  # 找到以.txt结尾的文件
        txt_path_list.append(file)
    elif file.endswith('.cs') or file.endswith('.json') or file.endswith('.xml'):
        cs_path_list.append(file)


def find_all_case(path):
    str_path = str(path)
    dir_files = os.listdir(str_path)  # 找到当前文件夹下所有文件
    for case in dir_files:
        case_path = os.path.join(str_path, case)  # 拼接成绝对路径
        if os.path.isfile(case_path):  # 如果是文件,得到文件路径
            txt_or_cs(case_path)
        elif os.path.isdir(case_path):  # 如果是子文件夹,继续递归
            find_all_case(case_path)


def unique_PD():
    for x in patient_names:
        if x in dicom_names:
            duplicated_PD.append(x)
    # duplicated_PD = ("4DLung",
    #                  "BigAlFFS",
    #                  "BigAlHFS",
    #                  "Bill",
    #                  "CTSdemoProst",
    #                  "FFPabdo",
    #                  "FusionProstate",
    #                  "Liver",
    #                  "MRIAdaptFFP",
    #                  "MRIAdaptFFS",
    #                  "MRIAdaptHFP",
    #                  "MRIAdaptHFS",
    #                  "4DCIRS1",
    #                  "PatientLiver",
    #                  "ProtonHSmall") # Patient 和 dicom的重名

    for p in patient_names:
        for pd in duplicated_PD:
            if p == pd:
                break
        else:
            unique_patient.append(p)

    for d in dicom_names:
        for pd in duplicated_PD:
            if d == pd:
                break
        else:
            unique_dicom.append(d)


def makeDiagram():
    if os.path.exists(text_zuotu.get()):
        if selected_diagram_radio.get() == 1:
            Chart.Pie_Chart(readExcel())
        elif selected_diagram_radio.get() == 2:
            Chart.Bar_Chart(readExcel())
        elif selected_diagram_radio.get() == 3:
            Chart.H_Bar_Chart(readExcel())
        else:
            tk.messagebox.showinfo('提示', '请选择图表类型')
    else:
        tk.messagebox.showinfo('提示', '请选择有效的Excel文件')


def exportToExcel():
    if os.path.exists(text_zhuaqu.get()):
        find_all_case(selected_path.get())
        if selected_T_or_S_radio.get() == 1:
            if selected_P_or_D_radio.get() == 1:
                txt_PD_Export('patient')
            elif selected_P_or_D_radio.get() == 2:
                txt_PD_Export('dicom')
            else:
                tk.messagebox.showinfo('提示', '请选择Patient or Dicom')
        elif selected_T_or_S_radio.get() == 2:
            if selected_P_or_D_radio.get() == 1:
                cs_PD_Export('patient')
            elif selected_P_or_D_radio.get() == 2:
                cs_PD_Export('dicom')
            else:
                tk.messagebox.showinfo('提示', '请选择Patient or Dicom')
        else:
            tk.messagebox.showinfo('提示', '请选择Test or Scripting')
    else:
        tk.messagebox.showinfo('提示', '请选择有效的文件夹路径')


def selectExcel():
    selected_file.set('')
    file_name = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx')])

    text_zuotu.delete(0, END)  # entry里选择第二个的时候, 删除前一个
    text_zuotu.insert(INSERT, file_name)

    if file_name:
        selected_file.set(file_name)


def selectPath():
    selected_path.set('')
    target_path = filedialog.askdirectory(title='选择文件夹路径')

    text_zhuaqu.delete(0, END)  # entry里选择第二个的时候, 删除前一个
    text_zhuaqu.insert(INSERT, target_path)

    if target_path:
        selected_path.set(target_path)


def closeThisWindow():
    if messagebox.askokcancel("退出", "确定要退出吗?"):
        root.destroy()


def readExcel():
    # tk.messagebox.showinfo('提示', '处理Excel文件的示例程序。')
    cells_list.clear()
    PD_list.clear()
    path = selected_file.get()
    workbook = openpyxl.load_workbook(path)  # 获取文件

    sheet_names = workbook.sheetnames  # 获取所有表的名字
    sheet = workbook.active  # 获取活跃表对象
    row_num = sheet.max_row  # 获取表中最大的行数
    col_num = sheet.max_column  # 获取表中最大的列数

    for col in range(1, col_num + 1):  # 获取表中所有单元格 放入list
        for row in range(1, row_num + 1):
            cell = sheet.cell(row, col)
            # 排除空单元格以及 case名中可能存在的 patient 或 dicom
            if bool(re.search(r'TS-', cell.value)) is False and cell.value is not None:
                cells_list.append(cell.value)
    # patient_names = ("4DLung",
    #                  "30x30x30",
    #                  "0505new",
    #                  "0505new2",
    #                  "002441",
    #                  "002443",
    #                  "002443TRN",
    #                  "002445",
    #                  "002445TRN",
    #                  "123456AA",
    #                  "141414",
    #                  "306941",
    #                  "0009999012",
    #                  "0010010777",
    #                  "0010117535",
    #                  "20140130",
    #                  "20160783",
    #                  "20190814B",
    #                  "111222333",
    #                  "111222333_NoContours",
    #                  "333222111",
    #                  "666555444",
    #                  "999888777",
    #                  "AAMDRSS2016PS",
    #                  "AdaptCBCT",
    #                  "AdaptCBCT2",
    #                  "AdaptCTDemo",
    #                  "ADTProstate",
    #                  "ANON76497",
    #                  "ArcCheck",
    #                  "ATP13",
    #                  "AutoCP",
    #                  "AutoMargin51102",
    #                  "BigAlFFS",
    #                  "BigAlHFS",
    #                  "BigPhantom",
    #                  "Bill",
    #                  "BOLUS11",
    #                  "Brain5TargetsTRN",
    #                  "breast1",
    #                  "breast1IMRT",
    #                  "BreastTRN",
    #                  "c1htth",
    #                  "CarbonLung",
    #                  "cem0tgdznno",
    #                  "CHESTnonPHANTOM",
    #                  "closedCentralL",
    #                  "cn1xt",
    #                  "CNNS1",
    #                  "cPatient",
    #                  "cPatient.AlgTest",
    #                  "cPatientProton",
    #                  "CrossPhantom",
    #                  "CT1",
    #                  "ctMRdosePhan",
    #                  "CTMRX30",
    #                  "CTMRXM20Y30Z50",
    #                  "CTMRYM40",
    #                  "CTMRZM50",
    #                  "CTSdemoHN",
    #                  "CTSdemoProst",
    #                  "CTSlung1",
    #                  "CTSprostate1",
    #                  "CTwithCTandPET",
    #                  "CUBE120kV",
    #                  "cwkjc5m51",
    #                  "cxrrbvw3",
    #                  "DemoPatient11",
    #                  "DepthSSDtest",
    #                  "DEVIL",
    #                  "DVHPatient1",
    #                  "DVHStats",
    #                  "DynMinLeafGap",
    #                  "EKK999999",
    #                  "ePHANTOM",
    #                  "FFP0CHEST",
    #                  "FFP88811X",
    #                  "FFPabdo",
    #                  "FFPDCMORIGIN",
    #                  "FFPEDFeetToHea",
    #                  "FFPEDHeadToFee",
    #                  "FFS15",
    #                  "FFS88811X",
    #                  "FFSDCMORIGIN",
    #                  "FFSEDFeetToHea",
    #                  "FFSEDHeadToFee",
    #                  "FFSPELVIS",
    #                  "ForceEDpat",
    #                  "ForceFillHN",
    #                  "FrozenDose4",
    #                  "FrozenTest",
    #                  "FrozenTest2",
    #                  "FrozenTest3",
    #                  "FrozenTestProton",
    #                  "FusionProstate",
    #                  "HalBreastXiO",
    #                  "HeadNeck_air",
    #                  "HeadProne",
    #                  "HFP00HEAD",
    #                  "HFP0CHEST",
    #                  "HFP88811X",
    #                  "HFPDCMORIGIN",
    #                  "HFPEDFeetToHea",
    #                  "HFPEDHeadToFee",
    #                  "HFPPELVIS",
    #                  "HFS00HEAD",
    #                  "HFS0CHEST",
    #                  "HFS30",
    #                  "HFS88811X",
    #                  "HFSDCMORIGIN",
    #                  "HFSEDFeetToHea",
    #                  "HFSEDHeadToFee",
    #                  "HFSPELVIS",
    #                  "HN",
    #                  "HN1",
    #                  "HNBig",
    #                  "HNnonPHANTOM",
    #                  "HNrtogIMRT",
    #                  "IMPTcase1",
    #                  "inhomtest",
    #                  "InvliChecksum",
    #                  "ionPINRTFP",
    #                  "ionPINRTFS",
    #                  "ionPINRTHP",
    #                  "ionPINRTHS",
    #                  "JamesCookMonac",
    #                  "K5dQtHx29P",
    #                  "larnumslices",
    #                  "Liver",
    #                  "LLungSBRT",
    #                  "LucyPhantom",
    #                  "Lung_air",
    #                  "lung_air_1",
    #                  "M511SSR",
    #                  "Matrixx",
    #                  "me010666x",
    #                  "monCOUCH",
    #                  "monDRPProstate",
    #                  "monDVHProstate",
    #                  "MonFP15B",
    #                  "monFP20B",
    #                  "monFP25B",
    #                  "monFusionProst",
    #                  "monFusionProstSAVE",
    #                  "monHN",
    #                  "monHPQC3121",
    #                  "monIMRTslabPhantm",
    #                  "monLung",
    #                  "monMLCBrain",
    #                  "monMLCProstate",
    #                  "monOROPHARYNX",
    #                  "monPhantom",
    #                  "monPlans",
    #                  "monPROSTATE",
    #                  "monPROSTATEexport",
    #                  "monPROSTATEmulti",
    #                  "monPROSTATEOLD",
    #                  "monPROSTATEproton",
    #                  "monSpherePhan",
    #                  "MRcntrSCO",
    #                  "MRIAdaptFFP",
    #                  "MRIAdaptFFS",
    #                  "MRIAdaptHFP",
    #                  "MRIAdaptHFS",
    #                  "MRIAdaptOrient",
    #                  "MRIDemo",
    #                  "MRITriangles",
    #                  "MRLCouch",
    #                  "MRX1Y1Z1",
    #                  "MRX20YM30Z40",
    #                  "MRX60",
    #                  "MRXM65Y55Z60",
    #                  "MRY50",
    #                  "MRZM60",
    #                  "MultipleStructures",
    #                  "nonphilips",
    #                  "OcuPro",
    #                  "OROPHARYNX",
    #                  "P20200925",
    #                  "PATIMG01",
    #                  "PHANTOM",
    #                  "PhShiftFFP50",
    #                  "PhShiftFFS50",
    #                  "PhShiftHFP50",
    #                  "PhShiftHFS50",
    #                  "PlanQuality",
    #                  "PointReg",
    #                  "PRETTYPROTONS",
    #                  "PROFFS18",
    #                  "PROHFP15",
    #                  "PROHFS13",
    #                  "prostate",
    #                  "Prostate_air",
    #                  "Prostate_norm",
    #                  "ProstateNodes",
    #                  "ProstShiftSAM",
    #                  "R95689791",
    #                  "referencePhant",
    #                  "RFC33982",
    #                  "RFC34976",
    #                  "RFC35090",
    #                  "roll",
    #                  "ROQS25",
    #                  "SBRTLung",
    #                  "SBRTSpine",
    #                  "SFUD_2",
    #                  "SFUD_3",
    #                  "SFUD1",
    #                  "smokePATIENT",
    #                  "smokePATIENT2",
    #                  "SPCIMGTEST",
    #                  "SpotLevel",
    #                  "SPTOriginalEUR",
    #                  "SSDTest2",
    #                  "Test1Proton",
    #                  "Test2Proton",
    #                  "Test9Proton",
    #                  "ThoracicCTPET",
    #                  "TolBreast",
    #                  "TPPC17A03PH",
    #                  "TwoBolus",
    #                  "UnityPatient1",
    #                  "UPMCLIV2",
    #                  "UPMCPANC1",
    #                  "UserAuth",
    #                  "VarianVMAT",
    #                  "wccccm",
    #                  "WeightLoss",
    #                  "XiOLung",
    #                  "EHU1412",
    #                  "4DCIRS1",
    #                  "111111",
    #                  "0000463294.anonymized",
    #                  "0000492064.anonymized",
    #                  "42482527",
    #                  "42482527_PerformanceTest",
    #                  "77378376",
    #                  "A4DCIRS1",
    #                  "cPatientCTRS_NonSquare",
    #                  "DUCKDAISY",
    #                  "HeadNeck_air_PerformanceTest",
    #                  "HP6834",
    #                  "Liver-4537",
    #                  "LockedLiver",
    #                  "lung_air_PerfomanceTest",
    #                  "MRIDemo177",
    #                  "MRIDemoCP",
    #                  "PatientLiver",
    #                  "ProtonCase",
    #                  "ProtonHBig",
    #                  "ProtonHSmall",
    #                  "ProtonPlanReview",
    #                  "ROQS25",
    #                  "smokePATIENT",
    #                  "smokePATIENT2",
    #                  "Test1",
    #                  "Test1IDD",
    #                  "Test3",
    #                  "Test3Air",
    #                  "Test4",
    #                  "Test4Water",
    #                  "Test6",
    #                  "Test8IDD",
    #                  "Test10",
    #                  "Test11RBE")
    # dicom_names = ("1ATLABDO_FFPMR_Day1",
    #                "1ATLABDO_FFPMR_Day2",
    #                "1ATLABDO_FFSMR_Day1",
    #                "1ATLABDO_FFSMR_Day2",
    #                "1ATLABDO_HFPMR_Day1",
    #                "1ATLABDO_HFPMR_Day2",
    #                "1ATLABDO_HFSMR_Day1",
    #                "1ATLABDO_HFSMR_Day2",
    #                "3D",
    #                "3DSTACarc",
    #                "4DCIRS1",
    #                "4DLung",
    #                "4011",
    #                "393456",
    #                "17112006",
    #                "6012247133",
    #                "AddonMLC",
    #                "Applicator",
    #                "BigAlFFS",
    #                "BigAlHFS",
    #                "Bill",
    #                "BlackPixels",
    #                "Block",
    #                "BO2",
    #                "BolusImportLiver",
    #                "BolusOncentra2",
    #                "BREAST0719",
    #                "BUG_6255_PhantomPlan",
    #                "Carbon_FFP",
    #                "Carbon_FFP_T_PTS",
    #                "Carbon_FFS",
    #                "Carbon_FFS_T_PTS",
    #                "Carbon_HFP",
    #                "Carbon_HFP_T_PTS",
    #                "Carbon_HFS",
    #                "Carbon_HFS_T_PTS",
    #                "ch2o1xr2rmFFP",
    #                "ch2o1xr2rmFFS",
    #                "ch2o1xr2rmHFP",
    #                "ch2o1xr2rmHFS",
    #                "Chordoma_2",
    #                "ConformalRT",
    #                "cPatientNonSquarePixelMR",
    #                "CS_CT_iBEAM_CouchTop",
    #                "CS_CT_MULTISERIES",
    #                "CS_CTMR_PICKER",
    #                "CS_MR_TILT",
    #                "CS_MRI_NON_SQR_PX",
    #                "CS_NEARLYAXIALMR",
    #                "CS_NOMERG_SIEMENS",
    #                "CS_PET_BLACK_HOLES",
    #                "CS_PET_GESLICE",
    #                "CS_PET_GEVOLUME",
    #                "CS_PETCT_GEDISC_JUN2002",
    #                "CS_RTD_ECLIPSE",
    #                "CS_RTD_ECLIPSE_IMPORT",
    #                "CS_RTD_KONRAD",
    #                "CS_RTD_ONCENTRAsep2005",
    #                "CS_RTD_TOMOHN",
    #                "CS_RTD_XIOIMRTARC",
    #                "CS_RTD_XIORTPLINK",
    #                "CS_RTP_ACQSIM_APRIL2005",
    #                "CS_RTP_ADVSIM_JUL2005",
    #                "CS_RTP_ONCENTRA_JUL2005",
    #                "CS_RTP_THERAPLAN",
    #                "CS_RTP_XIO_FORSSIMPORT",
    #                "CS_RTS_ACQSIM_NASO",
    #                "CS_RTS_COHERENCE",
    #                "CT_RTS_COHERENCE",
    #                "CT_UnEven",
    #                "CT_VF",
    #                "CTMR",
    #                "CTOR03",
    #                "CTPETSiemens",
    #                "CTSdemoProst",
    #                "DCAT_CDR_Bug_Data",
    #                "Demo_Phantom_1",
    #                "DemoPatient10",
    #                "dMLC",
    #                "DRPDepthTest",
    #                "DuplicateInstanceNumber",
    #                "DuplicateSliceLocation",
    #                "DuplicateSliceLocationNoInstanceNumber",
    #                "DynConformalArc",
    #                "ED_TestPhantoms",
    #                "EnergyError",
    #                "ErrorThrowout",
    #                "FEET1st^SUPINE_ABC-DEFGHIJKLMNOPQRSTUVWXYZ.abcdefghijklmnopqrstu",
    #                "FFP_PhantomSkull",
    #                "FFPabdo",
    #                "FFS",
    #                "FFS_RealPatient",
    #                "FFS_roll-15",
    #                "Fieldsplit",
    #                "FusionProstate",
    #                "GE_AdvSim_2field_BreastPlan",
    #                "GEPatient",
    #                "GESIM",
    #                "Grid9",
    #                "HeadSupine",
    #                "HFP",
    #                "HFP_1024Images",
    #                "HFP_roll-15",
    #                "HFS_RealPatientRTSS",
    #                "HFS_roll-15",
    #                "InvalidSnout",
    #                "InvalidSP",
    #                "ionRSThicknessError",
    #                "IOPnoTranswithoutMacro",
    #                "IOPxMACRO",
    #                "Liver",
    #                "mARC",
    #                "MMRO17A02EL_REG_Nucletron",
    #                "MMRO17A06EL_REG_Varian",
    #                "MMRO17A06RA_REG",
    #                "MMRO17A06RA_REG_RaySearch",
    #                "MMRO17A10VA_REG_3SET",
    #                "MMRO19A03EL",
    #                "Mon4TESTING01",
    #                "MONACODVH",
    #                "MonIonDicom",
    #                "monPhan",
    #                "MR",
    #                "MR_OBLIQUE_DATA",
    #                "MR_PLANE_CUTS",
    #                "MRIAdaptFFP",
    #                "MRIAdaptFFS",
    #                "MRIAdaptHFP",
    #                "MRIAdaptHFP_IPP",
    #                "MRIAdaptHFS",
    #                "MRIAdaptHFS_IPP",
    #                "MRMultiPlan",
    #                "MRPuppetOrient",
    #                "MRTestDemo",
    #                "MRTestPatientHN",
    #                "MultiDelivery",
    #                "MultiDuplicateSliceLocation",
    #                "MultiFrac",
    #                "NomappingTU",
    #                "Oncentra3",
    #                "OrientationTestPatient1",
    #                "p5MUeeYS0w",
    #                "PatientLiver",
    #                "Ph.Orig.FFP",
    #                "Ph.Orig.FFS",
    #                "Ph.Orig.HFP",
    #                "Ph.Orig.HFS",
    #                "Ph.PrismFadeZ",
    #                "Ph.Shift.FFP",
    #                "Ph.Shift.FFS",
    #                "Ph.Shift.HFP",
    #                "Ph.Shift.HFS",
    #                "Ph.Shift.HFS_Simplify",
    #                "Phantom_NullStudy",
    #                "PhDosePrism10",
    #                "PhilipsPatient",
    #                "PhillipsPetCT",
    #                "ProneData",
    #                "prostate-umb1-CBCT",
    #                "prostate-umb1-MRI",
    #                "ProstShift_DICOM",
    #                "ProtonHSmall",
    #                "Regression_5.5",
    #                "RFC27418Dcm",
    #                "RolledCT",
    #                "RSIDerror",
    #                "SE000001anon",
    #                "secondarymonPROSTATE",
    #                "secondaryMRIDemo",
    #                "Segerror",
    #                "SetupBeam",
    #                "SFUD4_ClinicalCase",
    #                "SiemensPatient",
    #                "SnoutIDError",
    #                "SpecExprt",
    #                "SpecImg",
    #                "SPHERE120kV",
    #                "SSDtest",
    #                "SSDtest2",
    #                "stepnshoot",
    #                "SwedishDicom",
    #                "T1Prostate",
    #                "TestProstatePatient",
    #                "TOSHIBA_IMG",
    #                "TPPC17A03BR_REG_MIM",
    #                "TPPC17A04BR_REG_Brainlab",
    #                "TuneIDerror",
    #                "Unity01",
    #                "vmat",
    #                "vmatnew",
    #                "VSADtest",
    #                "Wedge",
    #                "WinLevelPET",
    #                "YawPitchRollCT",
    #                "ZeroSliceThickness",
    #                "ZZZ_CT_1.25mmThinkness",
    #                "4DCIRS1SRO",
    #                "4DElektaXVi",
    #                "4DElektaXviNoStudy",
    #                "4DGE1",
    #                "4DGE2",
    #                "4DGE4",
    #                "4DGE5",
    #                "4DGE6",
    #                "4DManyStudies",
    #                "4DPhilips1",
    #                "4DPhilips2",
    #                "4DPhilips3",
    #                "4DPhilips4",
    #                "4DPhilips5",
    #                "4DPhilips6",
    #                "4DPhilips7",
    #                "4DPhilips8",
    #                "4DSiemens1",
    #                "4DSiemens2",
    #                "4DSiemens3",
    #                "4DSiemens4",
    #                "4DSiemens5",
    #                "4DSiemens6",
    #                "4DToshiba4",
    #                "000006+1485",
    #                "000006+1541",
    #                "000006+1648",
    #                "0000061485",
    #                "CarpePTS",
    #                "crjcr1q",
    #                "CTdisney",
    #                "ctMRdosePhan_MRI",
    #                "ctMRdosePhanMRsurvey",
    #                "DataBug15",
    #                "DosePhantom",
    #                "DosePhantoms",
    #                "DuplicateSliceLocation4D",
    #                "ExcCase1",
    #                "I18Neiou",
    #                "IPPShift",
    #                "Multiple_Size_Image",
    #                "PET_Phantom",
    #                "PETSUVA",
    #                "PETSUVB",
    #                "phantom",
    #                "SamplePatientsForImport",
    #                "testSUV",
    #                "WINLEVEL_PRESET_3",
    #                "WINLEVEL_PRESET_4")

    for c in cells_list:
        for PD in all_patient_dicom:
            if re.search(r'\b%s\b' % PD, str(c), re.IGNORECASE):
                PD_list.append(PD)

        # for dicom in dicom_names:
        #     if re.search(r'\b%s\b' % dicom, str(c), re.IGNORECASE):
        #         PD_list.append(dicom)

    # 对list进行筛选
    total = {}  # 字典 包含所有的id和它的数量
    for item in PD_list:  # 找到每个patient_id的使用次数
        if item in total.keys():
            total[item] += 1
        else:
            total[item] = 1

    new_filter_num = combo_num.get()
    if re.match(r'^[0-9]\d*$', new_filter_num):
        if combo_symbol.get() == '>=':
            changed = {k: v for k, v in total.items() if v >= int(new_filter_num)}
            return changed
        elif combo_symbol.get() == '>':
            changed = {k: v for k, v in total.items() if v > int(new_filter_num)}
            return changed
        elif combo_symbol.get() == '<':
            changed = {k: v for k, v in total.items() if v < int(new_filter_num)}
            return changed
        elif combo_symbol.get() == '<=':
            changed = {k: v for k, v in total.items() if v <= int(new_filter_num)}
            return changed
        else:
            tk.messagebox.showinfo('提示', '请重新选择过滤条件')
            raise TypeError("%s is an invalid symbol value" % combo_symbol.get())
    else:
        tk.messagebox.showinfo('提示', '请重新选择过滤条件')
        raise TypeError("%s is an invalid int value" % new_filter_num)


def txt_PD_Export(sss):
    xlsx_listP = []
    xlsx_listD = []
    case_40_list = []
    for line in txt_path_list:  # 每一行都是一个TC的绝对地址
        address = line.rstrip('\n')  # 去掉每行末尾的换行符'\n'
        TS_name = address.split('\\')[-1].rstrip('.txt')  # 得到TC的名字

        f = open(address, encoding='utf-8')  # 打开地址下的TC errors='ignore'
        content = f.readlines()
        # print(content)

        for i in range(len(content)):
            if re.search(r'\bTest Cases\b', content[i], re.IGNORECASE):
                # print(content[i])  # i是 "*** Test Cases ***" 所在的行数
                if i - 20 >= 0:  # 切片超出范围会从后往前数
                    aa = content[(i - 20): (i + 20)]
                else:
                    aa = content[: (i + 20)]
                aa.insert(0, TS_name)  # 把case名字插入到第一个
                case_40_list.append(aa)

    # print(case_40_list)
    # print(len(case_40_list)) #aaa里是每个case从 "*** Test Cases ***" 开始往下取30行的列表组成的列表嵌套

    if sss == 'patient':
        for case_40 in case_40_list:  # 每个case的40行
            str_40 = str(case_40)
            # CT1, prostate 误差较大
            for case_1 in case_40:
                if re.search(r'\bCT1\b', case_1, re.IGNORECASE) and \
                        re.search(r'\bLoad patient\b', case_1, re.IGNORECASE):
                    np = [case_40[0], 'CT1']
                    if np not in xlsx_listP:
                        xlsx_listP.append(np)
                elif re.search(r'\bprostate\b', case_1, re.IGNORECASE) and \
                        re.search(r'\bLoad patient\b', case_1, re.IGNORECASE):
                    np = [case_40[0], 'prostate']
                    if np not in xlsx_listP:
                        xlsx_listP.append(np)
            # 搜索重名的patient和dicom
            for PD in duplicated_PD:
                if re.search(r'\b%s\b' % PD, str_40, re.IGNORECASE) and \
                        re.search(r'\bLoad patient\b', str_40, re.IGNORECASE):
                    np = [case_40[0], PD]
                    if np not in xlsx_listP:
                        xlsx_listP.append(np)
            # 搜索patient
            for p in unique_patient:
                if re.search(r'\b%s\b' % p, str_40, re.IGNORECASE) and \
                        re.search(r'\bLoad patient\b', str_40, re.IGNORECASE):
                    np = [case_40[0], p]
                    if np not in xlsx_listP:
                        xlsx_listP.append(np)
        xlsx_listP = sorted(xlsx_listP, key=lambda x: x[0], reverse=False)

        try:
            wb = Workbook()
            ws = wb.active  # 对第一张sheet进行操作
            ws.column_dimensions["A"].width = 70
            ws.column_dimensions["B"].width = 20
            cell1 = ws["A1"]
            cell1.value = "TS-names"
            cell1.fill = PatternFill("solid", fgColor="B0C4DE")
            cell2 = ws["B1"]
            cell2.value = "Patient"
            cell2.fill = PatternFill("solid", fgColor="B0C4DE")
            for row in range(len(xlsx_listP)):
                ws.append(xlsx_listP[row])
            wb.save(r"%s.xlsx" % entry_excel_title.get())
        except:
            tk.messagebox.showinfo('提示', 'Excel打开时无法导入, 请关闭Excel或更改Excel标题')

    elif sss == 'dicom':
        for case_40 in case_40_list:  # 每个case的40行
            str_40 = str(case_40)

            # 3D, MR, HFP, FFS 误差较大
            for case_1 in case_40:
                if re.search(r'\b3D\b', case_1, re.IGNORECASE) and \
                        re.search(r'\bLoad dicom\b', case_1, re.IGNORECASE):
                    np = [case_40[0], '3D']
                    if np not in xlsx_listD:
                        xlsx_listD.append(np)
                elif re.search(r'\bMR\b', case_1, re.IGNORECASE) and \
                        re.search(r'\bLoad dicom\b', case_1, re.IGNORECASE):
                    np = [case_40[0], 'MR']
                    if np not in xlsx_listD:
                        xlsx_listD.append(np)
                elif re.search(r'\bHFP\b', case_1, re.IGNORECASE) and \
                        re.search(r'\bLoad dicom\b', case_1, re.IGNORECASE):
                    np = [case_40[0], 'HFP']
                    if np not in xlsx_listD:
                        xlsx_listD.append(np)
                elif re.search(r'\bFFS\b', case_1, re.IGNORECASE) and \
                        re.search(r'\bLoad dicom\b', case_1, re.IGNORECASE):
                    np = [case_40[0], 'FFS']
                    if np not in xlsx_listD:
                        xlsx_listD.append(np)

            # 搜索重名的patient和dicom
            for PD in duplicated_PD:
                if re.search(r'\b%s\b' % PD, str_40, re.IGNORECASE) and \
                        re.search(r'\bLoad dicom\b', str_40, re.IGNORECASE):
                    np = [case_40[0], PD]
                    if np not in xlsx_listD:
                        xlsx_listD.append(np)
            # 搜索dicom
            for d in unique_dicom:
                if re.search(r'\b%s\b' % d, str_40, re.IGNORECASE) and \
                        re.search(r'\bLoad dicom\b', str_40, re.IGNORECASE):
                    np = [case_40[0], d]
                    if np not in xlsx_listD:
                        xlsx_listD.append(np)
        xlsx_listD = sorted(xlsx_listD, key=lambda x: x[0], reverse=False)

        try:
            wb = Workbook()
            ws = wb.active  # 对第一张sheet进行操作
            ws.column_dimensions["A"].width = 70
            ws.column_dimensions["B"].width = 20
            cell1 = ws["A1"]
            cell1.value = "TS-names"
            cell1.fill = PatternFill("solid", fgColor="B0C4DE")
            cell2 = ws["B1"]
            cell2.value = "Dicom"
            cell2.fill = PatternFill("solid", fgColor="B0C4DE")
            for row in range(len(xlsx_listD)):
                ws.append(xlsx_listD[row])
            wb.save(r"%s.xlsx" % entry_excel_title.get())
        except:
            tk.messagebox.showinfo('提示', 'Excel打开时无法导入, 请关闭Excel或更改Excel标题')


def cs_PD_Export(xxx):
    xlsx_listP = []
    xlsx_listD = []
    if xxx == 'patient':
        for line in cs_path_list:  # 每一行都是一个TC的绝对地址
            address = line.rstrip('\n')  # 去掉每行末尾的换行符'\n'
            TS_name = address.split('\\')[-1].rstrip('.cs').rstrip('.json').rstrip('.xml')  # 得到TC的名字
            f = open(address, encoding='utf-8')  # 打开地址下的TC errors='ignore'
            content = f.readlines()
            for aline in content:  # case里其中一行
                # 搜索patient
                for p in patient_names:
                    if re.search(r'\b%s\b' % p, aline, re.IGNORECASE):
                        np = [TS_name, p]
                        if np not in xlsx_listP:
                            xlsx_listP.append(np)
        xlsx_listP = sorted(xlsx_listP, key=lambda x: x[0], reverse=False)

        try:
            wb = Workbook()
            ws = wb.active  # 对第一张sheet进行操作
            ws.column_dimensions["A"].width = 70
            ws.column_dimensions["B"].width = 20
            cell1 = ws["A1"]
            cell1.value = "TS-names"
            cell1.fill = PatternFill("solid", fgColor="B0C4DE")
            cell2 = ws["B1"]
            cell2.value = "Patient"
            cell2.fill = PatternFill("solid", fgColor="B0C4DE")
            for row in range(len(xlsx_listP)):
                ws.append(xlsx_listP[row])
            wb.save(r"%s.xlsx" % entry_excel_title.get())
        except:
            tk.messagebox.showinfo('提示', 'Excel打开时无法导入, 请关闭Excel或更改Excel标题')

    elif xxx == 'dicom':
        for line in cs_path_list:  # 每一行都是一个TC的绝对地址
            address = line.rstrip('\n')  # 去掉每行末尾的换行符'\n'
            TS_name = address.split('\\')[-1].rstrip('.cs').rstrip('.json').rstrip('.xml')  # 得到TC的名字
            f = open(address, encoding='utf-8')  # 打开地址下的TC errors='ignore'
            content = f.readlines()
            for aline in content:  # case里其中一行
                # 搜索dicom
                for d in dicom_names:
                    if re.search(r'\b%s\b' % d, aline, re.IGNORECASE):
                        np = [TS_name, d]
                        if np not in xlsx_listD:
                            xlsx_listD.append(np)
        xlsx_listD = sorted(xlsx_listD, key=lambda x: x[0], reverse=False)

        try:
            wb = Workbook()
            ws = wb.active  # 对第一张sheet进行操作
            ws.column_dimensions["A"].width = 70
            ws.column_dimensions["B"].width = 20
            cell1 = ws["A1"]
            cell1.value = "TS-names"
            cell1.fill = PatternFill("solid", fgColor="B0C4DE")
            cell2 = ws["B1"]
            cell2.value = "Dicom"
            cell2.fill = PatternFill("solid", fgColor="B0C4DE")
            for row in range(len(xlsx_listD)):
                ws.append(xlsx_listD[row])
            wb.save(r"%s.xlsx" % entry_excel_title.get())
        except:
            tk.messagebox.showinfo('提示', 'Excel打开时无法导入, 请关闭Excel或更改Excel标题')


if __name__ == '__main__':
    root = Tk()
    # root.overrideredirect(True)  # 去掉边框
    # ctypes.windll.shcore.SetProcessDpiAwareness(1)
    # ScaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0)  # 调用api获得当前的缩放因子
    # root.tk.call('tk', 'scaling', ScaleFactor / 75)  # 设置缩放因子

    rootWidth = 700
    rootHeight = 400
    screenWidth = root.winfo_screenwidth()
    screenHeight = root.winfo_screenheight()
    x = int((screenWidth - rootWidth) / 2)
    y = int((screenHeight - rootHeight) / 2)
    root.title('PD_Diagram_Maker V1.1')
    root.geometry("%sx%s+%s+%s" % (rootWidth, rootHeight, x, y))
    # root.geometry("500x350+10+10")
    root.resizable(False, False)  # 长度高度不可变

    patient_names = ("CTPETwithDose",
                     "000000444",
                     "4DLung",
                     "30x30x30",
                     "0505new",
                     "0505new2",
                     "002441",
                     "002443",
                     "002443TRN",
                     "002445",
                     "002445TRN",
                     "123456AA",
                     "141414",
                     "306941",
                     "0009999012",
                     "0010010777",
                     "0010117535",
                     "20140130",
                     "20160783",
                     "20190814B",
                     "111222333",
                     "111222333_NoContours",
                     "333222111",
                     "666555444",
                     "999888777",
                     "AAMDRSS2016PS",
                     "AdaptCBCT",
                     "AdaptCBCT2",
                     "AdaptCTDemo",
                     "ADTProstate",
                     "ANON76497",
                     "ArcCheck",
                     "ATP13",
                     "AutoCP",
                     "AutoMargin51102",
                     "BigAlFFS",
                     "BigAlHFS",
                     "BigPhantom",
                     "Bill",
                     "BOLUS11",
                     "Brain5TargetsTRN",
                     "breast1",
                     "breast1IMRT",
                     "BreastTRN",
                     "c1htth",
                     "CarbonLung",
                     "cem0tgdznno",
                     "CHESTnonPHANTOM",
                     "closedCentralL",
                     "cn1xt",
                     "CNNS1",
                     "cPatient",
                     "cPatient.AlgTest",
                     "cPatientProton",
                     "CrossPhantom",
                     # "CT1",
                     "ctMRdosePhan",
                     "CTMRX30",
                     "CTMRXM20Y30Z50",
                     "CTMRYM40",
                     "CTMRZM50",
                     "CTSdemoHN",
                     "CTSdemoProst",
                     "CTSlung1",
                     "CTSprostate1",
                     "CTwithCTandPET",
                     "CUBE120kV",
                     "cwkjc5m51",
                     "cxrrbvw3",
                     "DemoPatient11",
                     "DepthSSDtest",
                     "DEVIL",
                     "DVHPatient1",
                     "DVHStats",
                     "DynMinLeafGap",
                     "EKK999999",
                     "ePHANTOM",
                     "FFP0CHEST",
                     "FFP88811X",
                     "FFPabdo",
                     "FFPDCMORIGIN",
                     "FFPEDFeetToHea",
                     "FFPEDHeadToFee",
                     "FFS15",
                     "FFS88811X",
                     "FFSDCMORIGIN",
                     "FFSEDFeetToHea",
                     "FFSEDHeadToFee",
                     "FFSPELVIS",
                     "ForceEDpat",
                     "ForceFillHN",
                     "FrozenDose4",
                     "FrozenTest",
                     "FrozenTest2",
                     "FrozenTest3",
                     "FrozenTestProton",
                     "FusionProstate",
                     "HalBreastXiO",
                     "HeadNeck_air",
                     "HeadProne",
                     "HFP00HEAD",
                     "HFP0CHEST",
                     "HFP88811X",
                     "HFPDCMORIGIN",
                     "HFPEDFeetToHea",
                     "HFPEDHeadToFee",
                     "HFPPELVIS",
                     "HFS00HEAD",
                     "HFS0CHEST",
                     "HFS30",
                     "HFS88811X",
                     "HFSDCMORIGIN",
                     "HFSEDFeetToHea",
                     "HFSEDHeadToFee",
                     "HFSPELVIS",
                     "HN",
                     "HN1",
                     "HNBig",
                     "HNnonPHANTOM",
                     "HNrtogIMRT",
                     "IMPTcase1",
                     "inhomtest",
                     "InvliChecksum",
                     "ionPINRTFP",
                     "ionPINRTFS",
                     "ionPINRTHP",
                     "ionPINRTHS",
                     "JamesCookMonac",
                     "K5dQtHx29P",
                     "larnumslices",
                     "Liver",
                     "LLungSBRT",
                     "LucyPhantom",
                     "Lung_air",
                     "lung_air_1",
                     "M511SSR",
                     "Matrixx",
                     "me010666x",
                     "monCOUCH",
                     "monDRPProstate",
                     "monDVHProstate",
                     "MonFP15B",
                     "monFP20B",
                     "monFP25B",
                     "monFusionProst",
                     "monFusionProstSAVE",
                     "monHN",
                     "monHPQC3121",
                     "monIMRTslabPhantm",
                     "monLung",
                     "monMLCBrain",
                     "monMLCProstate",
                     "monOROPHARYNX",
                     "monPhantom",
                     "monPlans",
                     "monPROSTATE",
                     "monPROSTATEexport",
                     "monPROSTATEmulti",
                     "monPROSTATEOLD",
                     "monPROSTATEproton",
                     "monSpherePhan",
                     "MRcntrSCO",
                     "MRIAdaptFFP",
                     "MRIAdaptFFS",
                     "MRIAdaptHFP",
                     "MRIAdaptHFS",
                     "MRIAdaptOrient",
                     "MRIDemo",
                     "MRITriangles",
                     "MRLCouch",
                     "MRX1Y1Z1",
                     "MRX20YM30Z40",
                     "MRX60",
                     "MRXM65Y55Z60",
                     "MRY50",
                     "MRZM60",
                     "MultipleStructures",
                     "nonphilips",
                     "OcuPro",
                     "OROPHARYNX",
                     "P20200925",
                     "PATIMG01",
                     "PHANTOM",
                     "PhShiftFFP50",
                     "PhShiftFFS50",
                     "PhShiftHFP50",
                     "PhShiftHFS50",
                     "PlanQuality",
                     "PointReg",
                     "PRETTYPROTONS",
                     "PROFFS18",
                     "PROHFP15",
                     "PROHFS13",
                     # "prostate",
                     "Prostate_air",
                     "Prostate_norm",
                     "ProstateNodes",
                     "ProstShiftSAM",
                     "R95689791",
                     "referencePhant",
                     "RFC33982",
                     "RFC34976",
                     "RFC35090",
                     "roll",
                     "ROQS25",
                     "SBRTLung",
                     "SBRTSpine",
                     "SFUD_2",
                     "SFUD_3",
                     "SFUD1",
                     "smokePATIENT",
                     "smokePATIENT2",
                     "SPCIMGTEST",
                     "SpotLevel",
                     "SPTOriginalEUR",
                     "SSDTest2",
                     "Test1Proton",
                     "Test2Proton",
                     "Test9Proton",
                     "ThoracicCTPET",
                     "TolBreast",
                     "TPPC17A03PH",
                     "TwoBolus",
                     "UnityPatient1",
                     "UPMCLIV2",
                     "UPMCPANC1",
                     "UserAuth",
                     "VarianVMAT",
                     "wccccm",
                     "WeightLoss",
                     "XiOLung",
                     "EHU1412",
                     "4DCIRS1",
                     "111111",
                     "0000463294.anonymized",
                     "0000492064.anonymized",
                     "42482527",
                     "42482527_PerformanceTest",
                     "77378376",
                     "A4DCIRS1",
                     "cPatientCTRS_NonSquare",
                     "DUCKDAISY",
                     "HeadNeck_air_PerformanceTest",
                     "HP6834",
                     "Liver-4537",
                     "LockedLiver",
                     "lung_air_PerfomanceTest",
                     "MRIDemo177",
                     "MRIDemoCP",
                     "PatientLiver",
                     "ProtonCase",
                     "ProtonHBig",
                     "ProtonHSmall",
                     "ProtonPlanReview",
                     "ROQS25",
                     "smokePATIENT",
                     "smokePATIENT2",
                     # "Test1",
                     "Test1IDD",
                     # "Test3",
                     "Test3Air",
                     "Test4",
                     "Test4Water",
                     "Test6",
                     "Test8IDD",
                     "Test10",
                     "Test11RBE")  # 完整的patient list, 除了 # "CT1", "prostate", "Test1", "Test3", 误差较大
    dicom_names = ("1ATLABDO_FFPMR_Day1",
                   "1ATLABDO_FFPMR_Day2",
                   "1ATLABDO_FFSMR_Day1",
                   "1ATLABDO_FFSMR_Day2",
                   "1ATLABDO_HFPMR_Day1",
                   "1ATLABDO_HFPMR_Day2",
                   "1ATLABDO_HFSMR_Day1",
                   "1ATLABDO_HFSMR_Day2",
                   # "3D",
                   "3DSTACarc",
                   "4DCIRS1",
                   "4DLung",
                   "4011",
                   "393456",
                   "17112006",
                   "6012247133",
                   "AddonMLC",
                   "Applicator",
                   "BigAlFFS",
                   "BigAlHFS",
                   "Bill",
                   "BlackPixels",
                   "Block",
                   "BO2",
                   "BolusImportLiver",
                   "BolusOncentra2",
                   "BREAST0719",
                   "BUG_6255_PhantomPlan",
                   "Carbon_FFP",
                   "Carbon_FFP_T_PTS",
                   "Carbon_FFS",
                   "Carbon_FFS_T_PTS",
                   "Carbon_HFP",
                   "Carbon_HFP_T_PTS",
                   "Carbon_HFS",
                   "Carbon_HFS_T_PTS",
                   "ch2o1xr2rmFFP",
                   "ch2o1xr2rmFFS",
                   "ch2o1xr2rmHFP",
                   "ch2o1xr2rmHFS",
                   "Chordoma_2",
                   "ConformalRT",
                   "cPatientNonSquarePixelMR",
                   "CS_CT_iBEAM_CouchTop",
                   "CS_CT_MULTISERIES",
                   "CS_CTMR_PICKER",
                   "CS_MR_TILT",
                   "CS_MRI_NON_SQR_PX",
                   "CS_NEARLYAXIALMR",
                   "CS_NOMERG_SIEMENS",
                   "CS_PET_BLACK_HOLES",
                   "CS_PET_GESLICE",
                   "CS_PET_GEVOLUME",
                   "CS_PETCT_GEDISC_JUN2002",
                   "CS_RTD_ECLIPSE",
                   "CS_RTD_ECLIPSE_IMPORT",
                   "CS_RTD_KONRAD",
                   "CS_RTD_ONCENTRAsep2005",
                   "CS_RTD_TOMOHN",
                   "CS_RTD_XIOIMRTARC",
                   "CS_RTD_XIORTPLINK",
                   "CS_RTP_ACQSIM_APRIL2005",
                   "CS_RTP_ADVSIM_JUL2005",
                   "CS_RTP_ONCENTRA_JUL2005",
                   "CS_RTP_THERAPLAN",
                   "CS_RTP_XIO_FORSSIMPORT",
                   "CS_RTS_ACQSIM_NASO",
                   "CS_RTS_COHERENCE",
                   "CT_RTS_COHERENCE",
                   "CT_UnEven",
                   "CT_VF",
                   "CTMR",
                   "CTOR03",
                   "CTPETSiemens",
                   "CTSdemoProst",
                   "DCAT_CDR_Bug_Data",
                   "Demo_Phantom_1",
                   "DemoPatient10",
                   "dMLC",
                   "DRPDepthTest",
                   "DuplicateInstanceNumber",
                   "DuplicateSliceLocation",
                   "DuplicateSliceLocationNoInstanceNumber",
                   "DynConformalArc",
                   "ED_TestPhantoms",
                   "EnergyError",
                   "ErrorThrowout",
                   "FEET1st^SUPINE_ABC-DEFGHIJKLMNOPQRSTUVWXYZ.abcdefghijklmnopqrstu",
                   "FFP_PhantomSkull",
                   "FFPabdo",
                   # "FFS",
                   "FFS_RealPatient",
                   "FFS_roll-15",
                   "Fieldsplit",
                   "FusionProstate",
                   "GE_AdvSim_2field_BreastPlan",
                   "GEPatient",
                   "GESIM",
                   "Grid9",
                   "HeadSupine",
                   # "HFP",
                   "HFP_1024Images",
                   "HFP_roll-15",
                   "HFS_RealPatientRTSS",
                   "HFS_roll-15",
                   "InvalidSnout",
                   "InvalidSP",
                   "ionRSThicknessError",
                   "IOPnoTranswithoutMacro",
                   "IOPxMACRO",
                   "Liver",
                   "mARC",
                   "MMRO17A02EL_REG_Nucletron",
                   "MMRO17A06EL_REG_Varian",
                   "MMRO17A06RA_REG",
                   "MMRO17A06RA_REG_RaySearch",
                   "MMRO17A10VA_REG_3SET",
                   "MMRO19A03EL",
                   "Mon4TESTING01",
                   "MONACODVH",
                   "MonIonDicom",
                   "monPhan",
                   # "MR",
                   "MR_OBLIQUE_DATA",
                   "MR_PLANE_CUTS",
                   "MRIAdaptFFP",
                   "MRIAdaptFFS",
                   "MRIAdaptHFP",
                   "MRIAdaptHFP_IPP",
                   "MRIAdaptHFS",
                   "MRIAdaptHFS_IPP",
                   "MRMultiPlan",
                   "MRPuppetOrient",
                   "MRTestDemo",
                   "MRTestPatientHN",
                   "MultiDelivery",
                   "MultiDuplicateSliceLocation",
                   "MultiFrac",
                   "NomappingTU",
                   "Oncentra3",
                   "OrientationTestPatient1",
                   "p5MUeeYS0w",
                   "PatientLiver",
                   "Ph.Orig.FFP",
                   "Ph.Orig.FFS",
                   "Ph.Orig.HFP",
                   "Ph.Orig.HFS",
                   "Ph.PrismFadeZ",
                   "Ph.Shift.FFP",
                   "Ph.Shift.FFS",
                   "Ph.Shift.HFP",
                   "Ph.Shift.HFS",
                   "Ph.Shift.HFS_Simplify",
                   "Phantom_NullStudy",
                   "PhDosePrism10",
                   "PhilipsPatient",
                   "PhillipsPetCT",
                   "ProneData",
                   "prostate-umb1-CBCT",
                   "prostate-umb1-MRI",
                   "ProstShift_DICOM",
                   "ProtonHSmall",
                   "Regression_5.5",
                   "RFC27418Dcm",
                   "RolledCT",
                   "RSIDerror",
                   "SE000001anon",
                   "secondarymonPROSTATE",
                   "secondaryMRIDemo",
                   "Segerror",
                   "SetupBeam",
                   "SFUD4_ClinicalCase",
                   "SiemensPatient",
                   "SnoutIDError",
                   "SpecExprt",
                   "SpecImg",
                   "SPHERE120kV",
                   "SSDtest",
                   "SSDtest2",
                   "stepnshoot",
                   "SwedishDicom",
                   "T1Prostate",
                   "TestProstatePatient",
                   "TOSHIBA_IMG",
                   "TPPC17A03BR_REG_MIM",
                   "TPPC17A04BR_REG_Brainlab",
                   "TuneIDerror",
                   "Unity01",
                   "vmat",
                   "vmatnew",
                   "VSADtest",
                   "Wedge",
                   "WinLevelPET",
                   "YawPitchRollCT",
                   "ZeroSliceThickness",
                   "ZZZ_CT_1.25mmThinkness",
                   "4DCIRS1SRO",
                   "4DElektaXVi",
                   "4DElektaXviNoStudy",
                   "4DGE1",
                   "4DGE2",
                   "4DGE4",
                   "4DGE5",
                   "4DGE6",
                   "4DManyStudies",
                   "4DPhilips1",
                   "4DPhilips2",
                   "4DPhilips3",
                   "4DPhilips4",
                   "4DPhilips5",
                   "4DPhilips6",
                   "4DPhilips7",
                   "4DPhilips8",
                   "4DSiemens1",
                   "4DSiemens2",
                   "4DSiemens3",
                   "4DSiemens4",
                   "4DSiemens5",
                   "4DSiemens6",
                   "4DToshiba4",
                   "000006+1485",
                   "000006+1541",
                   "000006+1648",
                   "0000061485",
                   "CarpePTS",
                   "crjcr1q",
                   "CTdisney",
                   "ctMRdosePhan_MRI",
                   "ctMRdosePhanMRsurvey",
                   "DataBug15",
                   "DosePhantom",
                   "DosePhantoms",
                   "DuplicateSliceLocation4D",
                   "ExcCase1",
                   "I18Neiou",
                   "IPPShift",
                   "Multiple_Size_Image",
                   "PET_Phantom",
                   "PETSUVA",
                   "PETSUVB",
                   "phantom",
                   "SamplePatientsForImport",
                   "testSUV",
                   "WINLEVEL_PRESET_3",
                   "WINLEVEL_PRESET_4")

    duplicated_PD = []
    unique_patient = []
    unique_dicom = []
    unique_PD()
    all_patient_dicom = unique_patient + unique_dicom + duplicated_PD

    # tab1 variables
    txt_path_list = []
    cs_path_list = []
    selected_path = StringVar()
    selected_P_or_D_radio = IntVar()
    selected_T_or_S_radio = IntVar()

    # tab2 variables
    cells_list = []
    PD_list = []
    selected_diagram_radio = IntVar()
    selected_file = StringVar()
    selected_file.set('')
    filter_var = IntVar()
    filter_var.set('>=')  # 默认初始值为(>=0), 也就是不进行过滤
    filter_num = IntVar()
    filter_num.set('0')

    tab_main = ttk.Notebook()  # 创建分页栏
    tab_main.place(relx=0, rely=0, relwidth=1, relheight=1)

    tab1 = Frame(tab_main)  # 创建第一页框架

    tab1.place(x=0, y=30)
    tab_main.add(tab1, text='                                 抓取                                 ')  # 将第一页插入分页栏中
    label_zhuaqu = Label(tab1, text="请选择路径:")
    text_zhuaqu = Entry(tab1, background="lightgrey", width=42)
    button11 = Button(tab1, text="浏览", width=8, command=selectPath)  # 浏览
    button22 = Button(tab1, text="导出", width=8, command=exportToExcel)  # 导出
    button33 = Button(tab1, text="退出", width=8, command=closeThisWindow)  # 退出
    label_P_or_D = Label(tab1, text="抓取条件:")
    radio_test = Radiobutton(tab1, text="Test", value=1, variable=selected_T_or_S_radio)
    radio_scripting = Radiobutton(tab1, text="Scripting", value=2, variable=selected_T_or_S_radio)
    label___ = Label(tab1, text="-------------------------------")
    radio_patient = Radiobutton(tab1, text="Patient", value=1, variable=selected_P_or_D_radio)
    radio_dicom = Radiobutton(tab1, text="Dicom", value=2, variable=selected_P_or_D_radio)
    label_excel_title = Label(tab1, text="Excel标题:")
    entry_excel_title = Entry(tab1, width=20)
    entry_excel_title.insert(END, "Test_Data")

    # <editor-fold desc="抓取UI位置参数">
    label_zhuaqu.pack()
    text_zhuaqu.pack()
    button11.pack()
    button22.pack()
    button33.pack()
    label_P_or_D.pack()
    radio_patient.pack()
    radio_dicom.pack()
    label___.pack()
    radio_test.pack()
    radio_scripting.pack()
    label_excel_title.pack()
    entry_excel_title.pack()

    label_zhuaqu.place(x=25, y=40)
    text_zhuaqu.place(x=115, y=40)
    button11.place(x=500, y=38)
    button22.place(x=210, y=300)
    button33.place(x=310, y=300)
    label_P_or_D.place(x=170, y=80)
    radio_test.place(x=250, y=80)
    radio_scripting.place(x=350, y=80)
    label___.place(x=250, y=105)
    radio_patient.place(x=250, y=130)
    radio_dicom.place(x=350, y=130)
    label_excel_title.place(x=170, y=210)
    entry_excel_title.place(x=250, y=210)
    # </editor-fold>

    tab2 = Frame(tab_main)
    tab2.place(x=100, y=30)
    tab_main.add(tab2, text='                                 作图                                 ')
    label_zuotu = Label(tab2, text="请选择文件:")
    text_zuotu = Entry(tab2, background="lightgrey", width=42)
    button1 = Button(tab2, text="浏览", width=8, command=selectExcel)  # 浏览
    button2 = Button(tab2, text="作图", width=8, command=makeDiagram)  # 作图
    button3 = Button(tab2, text="退出", width=8, command=closeThisWindow)  # 退出
    radio_pie = Radiobutton(tab2, text="饼图", value=1, variable=selected_diagram_radio)
    radio_bar = Radiobutton(tab2, text="柱状图", value=2, variable=selected_diagram_radio)
    radio_Hbar = Radiobutton(tab2, text="横条图", value=3, variable=selected_diagram_radio)
    label_filter = Label(tab2, text="过滤条件:")
    combo_symbol = ttk.Combobox(tab2, textvariable=filter_var, width=3, values=['>', '>=', '<', '<='])
    combo_num = Entry(tab2, textvariable=filter_num, width=3)
    label_diagram_title = Label(tab2, text="图表标题:")
    label_x = Label(tab2, text="X轴名称:")
    label_y = Label(tab2, text="Y轴名称:")
    entry_diagram_title = Entry(tab2, width=20)
    entry_x = Entry(tab2, width=20)
    entry_y = Entry(tab2, width=20)

    # <editor-fold desc="作图UI位置参数">
    label_zuotu.pack()  # 调用pack方法将label标签显示在主界面
    text_zuotu.pack()
    button1.pack()
    button2.pack()
    button3.pack()
    radio_pie.pack()
    radio_bar.pack()
    radio_Hbar.pack()
    label_filter.pack()
    combo_symbol.pack()
    combo_num.pack()
    label_diagram_title.pack()
    label_x.pack()
    label_y.pack()
    entry_diagram_title.pack()
    entry_x.pack()
    entry_y.pack()

    label_zuotu.place(x=25, y=40)
    text_zuotu.place(x=115, y=40)
    button1.place(x=500, y=38)  # 浏览
    button2.place(x=210, y=300)  # 作图
    button3.place(x=310, y=300)  # 退出
    radio_pie.place(x=150, y=80)
    radio_bar.place(x=260, y=80)
    radio_Hbar.place(x=370, y=80)

    label_filter.place(x=170, y=130)
    combo_symbol.place(x=250, y=130)
    combo_num.place(x=310, y=130)

    label_diagram_title.place(x=170, y=170)
    label_x.place(x=170, y=210)
    label_y.place(x=170, y=250)
    entry_diagram_title.place(x=250, y=170)
    entry_x.place(x=250, y=210)
    entry_y.place(x=250, y=250)
    # </editor-fold>

    root.mainloop()
